import openpyxl

global data_file
data_set = []

try:
    data_file = openpyxl.load_workbook("TestData.xlsx")
except FileNotFoundError:
    data_file = openpyxl.load_workbook("TestData2.xlsx")

print(data_file)

sheet = data_file.active

# print(sheet.rows)
# print(type(next(sheet.rows)))

for rows in sheet.iter_rows()

    row_data = []
    for eachcell in rows:
        row_data.append(eachcell.value)

    print(row_data)
    data_set.append(row_data)

print(data_set)


# import xlrd

